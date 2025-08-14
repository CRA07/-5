import os
import re
import pandas as pd
from flask import Flask, request
import openpyxl
from datetime import datetime
from filelock import FileLock
from openpyxl.reader.excel import load_workbook

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "data", "popitka5.xlsx")


WEBHOOK_TOKEN = os.getenv("WEBHOOK_TOKEN", "token20220705")          # необязательный токен ?token=...
PORT = int(os.getenv("PORT", "8000"))
BIND_HOST = os.getenv("BIND_HOST", "0.0.0.0")

LOCK_FILE = EXCEL_FILE + ".lock"



# Ключевые слова для поиска
PRODUCTS = ["STZ_Agenta_Aжента_100", "PML_PML_Завтрак_200",	"PML_PML_Хлорофил_500",	"KSM_kosmoteros_СывВитКомп_30",	"KSM_Kosmoteros_ТоникНорм_200",	"KSM_Kosmoteros_ТоникСухая_200",	"KSM_Kosmoteros_ТоникЖирная_200",	"KSM_kosmoteros_ТоникВитС_200",	"KSM_ГельПенкаМат_150",	"KSM_ГельПенкаBG_150",	"KSM_kosmoteros_ГельAHA_200мл",	"KSM_kosmoteros_ГельВсеТипыКожи_200",	"KSM_kosmoteros_себорегулятор_50",	"KSM_kosmoteros_крем15spf_50",	"KSM_kosmoteros_КонцАнтиЭйджРевит_30",	"KSM_kosmoteros_КонцЛифт_30",	"KSM_kosmoteros_КонцСуперУвлаж_30",	"KSM_kosmoteros_КонцЭластКожа_30",	"KSM_kosmoteros_КонцГлик_30",	"KSM_kosmoteros_КремИкра_25",	"KSM_kosmoteros_КремSL_25",	"KSM_kosmoteros_КремВитС_50",	"KSM_kosmoteros_КремСПФ30_50",	"KSM_kosmoteros_Ремодулятор_30",	"KSM_kosmoteros_кремIB_50",	"KSM_kosmoteros_крембриллиант_50",	"KSM_Kosmoteros_МаскаШокУдов_10",	"KSM_Kosmoteros_МаскаЭнзимПил_10",	"KSM_Kosmoteros_МаскаНежныеСливки_50",	"KSM_Kosmoteros_МаскаВекиHYASEALON_1",	"KSM_Kosmoteros_МаскаЭкспрессЛифтинг_50",	"KSM_Kosmoteros_МаскаИммунокор_50",	"KSM_Kosmoteros_МаскаСеборег_50",	"KSM_Kosmoteros_МаскаBeauteGlobale_50",	"KSM_Kosmoteros_МаскаМорКоллаген_50",	"KSM_Kosmoteros_ПилингГоммаж_50",	"KSM_Kosmoteros_ЛосьонГидроактиватор_200",	"KSM_Kosmoteros_КремМорКоллаген_50",	"KSM_Kosmoteros_ЛосьонБиоактиватор_200",	"KSM_Kosmoteros_КремSuperGold_80",	"KSM_Kosmoteros_ГельAnticouperose_50",	"AFP_AFP_Ф8Псоризан_90",	"AFP_AFP_Ф9Куперозан_75",	"AFP_AFP_КремПапилФит_5",	"AFP_AFP_ЭкстрЛопух_60",	"ALF_Alfit_ФитосборМастопатия_60",	"ALF_Alfit_ФитосборГинекология_60",	"ALF_Alfit_ФитосвечиПростозол_15",	"ALF_Alfit_ФитосвечиГинекозол_15",	"ALF_Alfit_ФитосборДляСниженияВеса_60",	"ALF_Alfit_ФитосборДляЖКТ_60",	"ALF_Alfit_ФитосборДляМозговогоКровообращения_60",	"ALF_Alfit_ФитосборДляПечени_60",	"ALF_Alfit_ФитосборПротивопаразитный_60",	"ALF_Alfit_ФитосборДляНормализацииСахара_60",	"ALF_Alfit_ФитосборДляИммуннойСистемы_60",	"ALF_Alfit_ФитосборКлимактерический_60",	"ALF_Alfit_КапсулыГинеконорм_60",	"ALF_Alfit_КапсулыМастонорм_60",	"ALF_Alfit_ПанталфитМужской_30",	"ALF_Alfit_Лисички_30",	"ALF_Alfit_ЧагаРейше_30",	"FMT_Ecoderm_Шампунь_150",	"FMT_Regecin_Гель_15",	"FMT_Sebozol_Шампунь_100",	"FMT_Sebozol_Шампунь_200",	"VTB_Menopace_Менопейс_30",	"VTB_Wellman_Таблетки_30",	"VTB_Calcimax_Кальцимакс_60",	"VTB_Perfectil_Таблетки_30",	"VTB_Perfectil_Трихолоджик_60",	"VTB_Feroglobin_Фероглобин_30",	"VTB_Perfectil_Платинум_30",	"VTB_Wellman_Трихолоджик_60",	"VTB_Wellwoman_50+_30",	"VTB_Menopace_Плюс_28&28",	"VTB_Perfectil_Платинум_60",	"FMT_Sebozol_Шампунь_25",	"LTM_Vitaon_БальзамЛюкс_25",	"PLS_Admera_Крем_150",	"PLS_Admera_Крем_50",	"PLS_Отофаг_Гель_50",	"PLS_Суперчистотело_Маркер_5",	"UNL_Axe_ГифтПакШампГддДезDT_600",	"UNL_Axe_ГельШампуньФеникс_610",	"UNL_Axe_ГельШампуньANTI_HANGOVER_610",	"FMT_Ecoderm_Шампунь_750",	"SHB_Dr.Ohhira_Коллаген_10",	"SHB_Dr.Ohhira_Коллаген_15",	"SHB_Dr.Ohhira_РастительныйКомплекс_120",	"SHB_Dr.Ohhira_РастительныйКомплекс_60",	"SHB_Dr.Ohhira_РастительныйПреметабиотик_30",	"SHB_Dr.Ohhira_РастительныйКомплекс_30",	"UNL_Axe_ШампуньГельCoolOcean_610",	"UNL_Axe_ГельШампуньEPIC_FRESH_610",	"UNL_Axe_ГельШампуньНаПляже_610",	"UNL_Axe_ГельКожаПеченьки_610",	"UNL_Axe_ГельDarkTemptation_610",	"UNL_Axe_ГельIceChill_610",	"UNL_Axe_ГельШампуньАкватическийБергамот_400",	"UNL_Axe_ГельСилаДжунглей_610",	"UNL_Axe_ГельСилаТехнологий_610",	"UNL_Axe_ГельИзумрудныйПачули_400",	"UNL_Axe_ГельМедныйСандал _400",	"UNL_Axe_ГельШампуньЯнтарноеМанго_400",	"UNL_Axe_ДезодорантАкватическийБергамот_150",	"UNL_Axe_ДезодорантИзумрудныйПачули_150",	"UNL_Axe_ДезодорантМедныйСандал_150",	"UNL_Axe_ДезодорантЯнтарноеМанго_150",	"UNL_Axe_ДезСтикDarkT_50",	"UNL_Axe_НаборГддДезDT_250",	"UNL_Axe_ГифтУдача_400",	"UNL_Camay_ГельВосхитительноеОчарование_750",	"UNL_Camay_ГельГипнотическаяАура_750",	"UNL_Camay_ГельМагическоеЗаклинание_750",	"UNL_Camay_МистПионЯгоды_115",	"UNL_Camay_НаборГелей_2х250",	"UNL_Camay_НаборГддМочалка_250",	"UNL_Camay_ГельТайноеБлаженство_750",	"UNL_Cif_НаборАПАНАЖ_1500",	"UNL_Cif_СпрейАнтижир_500",	"UNL_Cif_СпрейКремПена_500",	"UNL_Clear_CDEШампуньДляРоста_380",	"UNL_Clear_CDEШампуньОтПерхоти_380",	"UNL_Clear_CDEСывороткаДляВолос_190",	"UNL_Clear_CDTМаскаКондЛОК_200",	"UNL_Clear_CDTМаскаКондКИУ_200",	"UNL_Clear_CDTШампуньМягкийКИУ_380",	"UNL_Clear_CDTШампуньОсвежЛОК_380",	"UNL_Clear_CDTСкрабДляВолосЭР_150",	"UNL_Clear_CDTСывороткаДляВолосЭР_190",	"UNL_Clear_CDTМаскаКондЭР_200",	"UNL_Clear_CDTШампуньЭР_380",	"UNL_Clear_ШапуньБальзамMen_610",	"UNL_Clear_ГельШампуньБальзамMen_610",	"UNL_Clear_ШампуньОтВыпаденияВолос_610",	"UNL_Clear_ПилингДляГоловы_150",	"UNL_CrystalPur_ГельОчищ_190",	"UNL_CrystalPur_ГидрофильноеМасло_110",	"UNL_Dove_КремКокос_300",	"UNL_Dove_КремГельПитаниеУвлажнение_610",	"UNL_Dove_КремГельИнжирАпельсин_610",	"UNL_Dove_КремГельФисташковыйКрем_610",	"UNL_Dove_НаборСухойШампунь_500",	"UNL_Dove_ГифтПакШампГдд_500",	"UNL_Dove_НаборКремДляРукДез_100    ",	"UNL_Dove_ШампБальзЛаванда_1260",	"UNL_Dove_ШампБальзМанго_1260",	"UNL_Dove_ШампБальзВосст_1260",	"UNL_Feel Moment_ГидроБальзам_45",	"UNL_Feel Moment_КремРетинол_50",	"UNL_Feel Moment_КремЦинкНиацин_50",	"UNL_Feel Moment_ПенкаДляЛица_190",	"UNL_Feel Moment_СывороткаВитС_30",	"UNL_Feel Moment_СывороткаРетинол_30",	"UNL_Rexona_ГддМужГорныйЛед_750",	"UNL_Rexona_НаборMenГддДез_230",	"UNL_Rexona_ГифтПакДезГдд23_330",	"UNL_Rexona_НаборГддДезПляж_240",	"UNL_Rexona_ДезКдрСердце_95",	"UNL_Tresemme_ШампуньБальзамКремВолны_880",	"UNL_Tresemme_БальзамКератин_930",	"UNL_Tresemme_ШампуньБальзамКератин_1930",	"UNL_Tresemme_ШампуньКератин_1000",	"UNL_Tresemme_МаскаДляВолосRep_300",	"UNL_Tresemme_МаскаДляВолосRichM_300",	"UNL_Tresemme_МаскаДляВолосОкраш_300",	"UNL_Tresemme_МаслоДляВолосRep_50",	"UNL_Tresemme_МуссДляВолос5Фикс_200",	"UNL_Tresemme_МуссДляВолос3Объем_200",	"FMT_AquaFoeniculi_УкропнаяВода_15",	"UNL_Tresemme_НаборСухойШампунь_500",	"UNL_Tresemme_НаборШампуньСпрей_420",	"UNL_Tresemme_СпрейДляВолос_190",	"UNL_Tresemme_СпрейДляВолосОбъемBFV_190",	"UNL_Tresemme_СывороткаФлюидRichM_190",	"UNL_Tresemme_ШампуньБальзам_1280",	"UNL_Tresemme_ШампуньБальзамОбъем_1280",	"UNL_Tresemme_ШампуньБальзамОкраш_1280",	"UNL_Tresemme_ШампуньБальзамУвлажнение_800",	"UNL_Tresemme_ШампуньБальзамДетокс_800",	"UNL_Tresemme_ШампуньБальзамБриллиант_800",	"UNL_Tresemme_ШампуньБальзамКератин_800",	"UNL_Tresemme_ШампуньБальзамRP_800",	"UNL_Tresemme_ШампуньБальзамОбъем_800",	"UNL_Tresemme_ШампуньБальзамОкраш_800",	"UNL_Tresemme_ШампуньБальзамВолны_800",	"UNL_Tresemme_ШампуньБальзамУвлажнение_1280",	"UNL_Бархатные ручки _ЖидкоеМылоПак_1000",	"UNL_Бархатные ручки _ЖидкоеКремМыло_5",	"UNL_Бархатные ручки_НаборКремов_90",	"UNL_Лесной бальзам_ЗПДозГранат_290",	"UNL_Лесной бальзам_ЗПДозМорскаяСоль_290",	"FMT_AquaFoeniculi_УкропнаяВодаСаше10",	"UNL_Лесной бальзам_ЗПОтбелХЗ_319",	"UNL_Лесной бальзам_ЗубнаяПастаTotal_290",	"UNL_Лесной бальзам_НаборСольИМятаДоз_580",	"UNL_Черный жемчуг_ЛосьонТоник_190",	"SHB_Dr.Ohhira_РастительныйКомплексDeLuxe_30",	"UNL_ЧернЖем_ПилингСкатка_120",	"UNL_Черный жемчуг_БазаПодМакияжСПФ_40",	"UNL_Черный жемчуг_ГельПилинг_80",	"UNL_Черный жемчуг_25+КремСПФ30_48",	"UNL_ЧернЖем_AntiAgeКремSPF30_48",	"UNL_Черный жемчуг_КремФиллерДневной_48",	"UNL_ЧернЖем_КремСкульптор_50",	"UNL_Черный жемчуг_КремДляВек_15",	"UNL_Черный жемчуг_МицеллВода_750",	"UNL_Черный жемчуг_МистПитание_115",	"UNL_Черный жемчуг_МистУвлажнение_115",	"UNL_Черный жемчуг_МультиКремБальз_50",	"UNL_Черный жемчуг_МультиКремДляВек_40",	"UNL_Черный жемчуг_СтикДляГуб_12",	"UNL_ЧернЖем_НочнКремМаскаLift_48",	"UNL_Черный жемчуг_ПилингМаскаОбнов_81",	"UNL_Черный жемчуг_СплэшМаска_160",	"UNL_Черный жемчуг_МультиСДСС_170",	"UNL_Черный жемчуг_ССКремСПФ25_40",	"UNL_Черный жемчуг_СывороткаБустерВосстанавл_30",	"UNL_Черный жемчуг_СывороткаБустер_30",	"UNL_Черный жемчуг_СывороткаБустерРазглаж_30",	"UNL_Черный жемчуг_СывороткаБустерУвлажн_30",	"UNL_Черный жемчуг_СывороткаБустерУкрепл_30",	"UNL_Черный жемчуг_СывороткаБустерУспокающ_30",	"UNL_Черный жемчуг_МультиКремТело_150",	"UNL_ЧистаяЛиния_ГельПерсик_750",	"UNL_ЧистаяЛиния_ГельПробуждающий_750",	"UNL_ЧистаяЛиния_ГельКлубника_750",	"UNL_ЧистаяЛиния_МицеллВода_750",	"VTB_Wellwoman_Плюс_28&28",	"VTB_Wellman_Плюс_28&28",	"VTB_Perfectil_Плюс_28&28",	"GRT_HealthIs_ЖелезоСульфат_60",	"GRT_HealthIs_ЛецитинПодсолн_120",	"GRT_Handy_МаслоМонарды_30",	"GRT_HealthIs_Куркумин_30",	"GRT_HealthIs_ДГК_30",	"GRT_HealthIs_Таурин_90",	"GRT_Handy_Ремувер_30",	"GRT_HealthIs_Биотин_60",	"GRT_HealthIs_Женьшень_60",	"GRT_HealthIs_Хром_60",	"GRT_HealthIs_Йод_60",	"GRT_HealthIs_КореньЛопуха_60",	"GRT_Handy_МЧД_30",	"GRT_HealthIs_АльфаЛипКислота_60",	"GRT_HealthIs_Тирозин_60",	"GRT_HealthIs_Д3_180",	"GRT_HealthIs_Глицин_60",	"GRT_HealthIs_Спирулина_90",	"GRT_HealthIs_Расторопша_1000_120",	"GRT_HealthIs_КоэнзимQ10_100_60",	"GRT_HealthIs_ЕжовикГребенчатый_2000_120",	"GRT_HealthIs_Селен_150_60",	"GRT_HealthIs_ФолиеваяКислота_600_180",	"GRT_HealthIs_Таурин_90",	"GRT_HealthIs_МагнийХелатВ6_120",	"GRT_HealthIs_В-Комплекс_60",	"GRT_Handy_КремОтРастяжек_250",	"GRT_HealthIs_Д3'5000_180",	"GRT_HealthIs_ЦинкПиколинат_120",	"GRT_HealthIs_ВитаминА_1000_90",	"GRT_HealthIs_ВитаминЕ_60",	"GRT_HealthIs_Хлорофилл_500",	"GRT_HealthIs_Берберин_60",	"GRT_HealthIs_ДетскиеВитамины_60",	"GRT_Handy_НаборМонардаРемувер_60",	"GRT_Handy_МаслоУсьмы_10",	"GRT_HealthIs_Астаксантин_90",	"GRT_HealthIs_Коллаген_90",	"GRT_HealthIs_Глицин_90",	"GRT_HealthIs_5-htp_60",	"GRT_HealthIs_Пробиотики_60",	"GRT_HealthIs_ZMA_120",	"GRT_HealthIs_КоллагенПорошок_30",	"GRT_HealthIs_ЯнтарнаяКислота_120",	"GRT_Handy_МаслоДляКутикулыИНогтей_30",	"GRT_Handy_МаслоДляЛицаМассажное_30",	"GRT_HealthIs_Омега3_180",	"GRT_HealthIs_Глюкозамин_90",	"GRT_HealthIs_МужскиеВитамины_60",	"GRT_HealthIs_ЖенскиеВитамины_60",	"GRT_HealthIs_ВитаминС_120",	"GRT_Handy_МассажноеМаслоДляТела_500",	"GRT_Handy_МаслоДляВолос_100",	"GRT_HealthIs_Инозитол_60",	"GRT_Handy_МаслоШи_150",	"GRT_Handy_МаслоРозмарина_30",	"GRT_HealthIs_Хром_120",	"GRT_HealthIs_Набор_4",	"GRT_HealthIs_Д3К2_120",	"GRT_Handy_МагниевоеМасло_200",	"GRT_HealthIs_ЖелезоХелат_60",	"GRT_HealthIs_Лецитин2000_120",	"KSM_kosmoteros_КремСПФ50_50",	"SHB_Dr.Ohhira_РастительныйКомплексDeLuxe_60",	"GRT_HealthIs_МагнийЦитрат_120",	"GRT_HealthIs_ГлицинДляДетей_60",	"GRT_HealthIs_ВитаминыДляВолос_60",	"UNL_ЧистаяЛиния_ГддКдрMen_330",	"UNL_Чистая Линия_НаборМульти_200",	"UNL_PureLine_СывороткаКоррект_30",	"UNL_PureLine_КремОтПрыщей_185",	"UNL_PureLine_КриоГель_45",	"UNL_PureLine_КриоПатчиДляВек_30",	"UNL_PureLine_ПенкаОтПрыщей_190",	"FMT_Regecin_НаборГель_30",	"GRT_HealthIs_Омега3_90",	"GRT_HealthIs_Омега369_180",	"UNL_CrystalPur_КремУвлажнЦерамиды_50",	"UNL_CrystalPur_СывороткаОмолаживающая_30",	"UNL_CrystalPur_СывороткаУвлажняющая_30",	"VTB_Pregnacare_Прегнакеар_28&28",	"GRT_HealthIs_К2_90",	"GRT_HealthIs_МаслоЧерногоТмина_90",	"UNL_Camay_ГельФранцузскаяРоза_750",	"UNL_Camay_ГельПенаЛавандаМагний_750",	"UNL_Camay_ГельЯпонскаяСакура_750",	"UNL_Camay_ГельРозовыйГрейпфрут_750",	"UNL_Camay_ГельВишняМиндаль_750",	"UNL_Camay_ГельГранатКоллаген_750",	"UNL_Camay_ГельЦитрусовыйМикс_750",	"UNL_Camay_ГельАбрикосМаслоШи_750",	"GRT_Handy_РемуверДляКутикулы_50",	"GRT_Handy_Крем-баттерДляТела_250",	"GRT_HealthIs_ЛипотропныйФактор_90",	"UNL_Dove_ШампБальзАвокадо_1260",	"UNL_Tresemme_НаборЛаков_500",	"UNL_ЧистаяЛиния_НаборПенок_300",	"UNL_Tresemme_СпрейДляВолосКератин_250",	"GRT_Healthis_МагнийЦитрат_240",	"GRT_HealthIs_МагнийХелатВ6_270",	"GRT_HealthIs_Глюкозамин_180",	"GRT_Kottur_МаслоДляЗагара_150",	"GRT_Kottur_КремСМочевиной_500",	"GRT_HealthIs_Креатин_300",	"GRT_HealthIs_Карнитин_90",	"GRT_Handy_МассажноеМаслоДляТела_500",	"GRT_HealthIs_Кальций_1000_120",	"GRT_Kottur_СПФДляЛица_50",	"GRT_HealthIs_Коллаген_180",	"GRT_HealthIs_Аргинин_180",	"GRT_HealthIs_Аргинин_90",]

WAREHOUSE_DEFECTS = ["пришел другой дозатор", "нет этикетки", "нет дозатора",
           "нет товара", "пришел разбитым", "перепутан штрихкод", "перепутан товар",
           "брак", "проблема с этикеткой", "просрочка", "нет упаковки"]

DEFECT_CATEGORIES = {
    "пришел другой дозатор": "комплектация",
    "нет этикетки": "этикетка",
    "нет дозатора": "комплектация",
    "нет товара": "комплектация",
    "пришел разбитым": "перевозка",
    "перепутан штрихкод": "этикетка",
    "перепутан товар": "смена товара",
    "брак": "брак",
    "проблема с этикеткой": "этикетка",
    "нет упаковки": "упаковка",
    "просрочка": "просрочка"
}
PRODUCTION_DEFECTS = ["нет даты производства ",	"волос в банке", "пустая банка", "банка без защитной фольги",	"расплавленный вид",	"недостающее количество капсул", "капсулы в масле", "капсулы в пятнах",	"черная крышка внутри банки",	"неприятный запах от капсул",	"пустые капсулы в банке ",	"просрочка",	"вскрытая банка",	"пришел без этикетки",	"вскрыт пакет с селикагелем",	"жалоба на плесень ",
]
MARKETPLACES = ["вб", "озон", "ям"]

def normalize(text: str) -> str:
    """Приводит текст к нижнему регистру, нормализует латиницу/кириллицу, убирает пробелы и подчёркивания"""
    text = text.lower()
    latin_to_cyrillic = {
        'a': 'а',
        'e': 'е',
        'o': 'о',
        'p': 'р',
        'c': 'с',
        'y': 'у',
        'x': 'х'
    }
    text = ''.join(latin_to_cyrillic.get(ch, ch) for ch in text)
    text = re.sub(r'[\s_]+', '', text)
    return text


def normalize(text):
    return re.sub(r'[\s_]+', '', text.lower())

def find_match(text, collection):
    text_norm = normalize(text)
    for item in collection:
        if normalize(item) in text_norm:
            return item
    return ""


def ensure_parent_dir(path: str):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)

# строка создания эксель файла
def init_excel():
    """Создаёт Excel, если нет."""
    ensure_parent_dir(EXCEL_FILE)
    if not os.path.exists(EXCEL_FILE):

        from openpyxl import Workbook
        filename = "popitka5.xlsx"
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Брак Склада"
        sheet1.append(["Дата", "Автор", "Код продукта", "Маркетплейс", "Описание проблемы", "Характеристика проблемы",
        "Текст сообщения"])
        sheet2 = wb.create_sheet(title="Производство")
        sheet2.append(["Дата", "Автор", "Код продукта", "Описание проблемы", "Текст сообщения"])
        wb.save(filename="popitka5.xlsx")


@app.route("/webhook", methods=["POST"])
def webhook():
    # --- Простая защита общим токеном (опционально) ---
    if WEBHOOK_TOKEN:
        if request.args.get("token") != WEBHOOK_TOKEN:
            return "", 403  # молча отклоняем

    data = request.get_json(silent=True) or {}
    # Формат Пачки (как в твоём примере)
    text = str(data.get("content", "")).lower()
    author_id = data.get("user_id", "Неизвестно")
    timestamp = data.get("created_at", datetime.now().isoformat())

    try:
        time_str = datetime.fromisoformat(str(timestamp).replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except:
        time_str = datetime.utcnow().strftime("%Y-%m-%d")


    def find_match(text, collection):
        for item in collection:
            if item.lower() in text:
                return item
        return ""

    if text.startswith("#склад"):
        content_without_tag = text.replace("#склад", "", 1).strip()

        found_product = find_match(content_without_tag, PRODUCTS)
        found_defect = find_match(content_without_tag, WAREHOUSE_DEFECTS)
        defect_category = DEFECT_CATEGORIES.get(found_defect, "") if found_defect else ""
        marketplace = find_match(content_without_tag, MARKETPLACES)

        # Пишем, даже если нет маркетплейса, но есть продукт и дефект
        if found_product and found_defect:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet = wb["Брак Склада"]
            sheet.append([
                time_str,
                author_id,
                found_product,
                marketplace if marketplace else "",
                found_defect,
                defect_category,
                text
            ])
            wb.save(EXCEL_FILE)

    elif text.startswith("#производство"):
        content_without_tag = text.replace("#производство", "", 1).strip()

        found_product = find_match(content_without_tag, PRODUCTS)
        found_complaint = find_match(content_without_tag, PRODUCTION_DEFECTS)

        if found_product and found_complaint:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet = wb["Производство"]
            sheet.append([
                time_str,
                author_id,
                found_product,
                found_complaint,
                text
            ])
            wb.save(EXCEL_FILE)

    return "", 200


if __name__ == "__main__":
    init_excel()
    app.run(host=BIND_HOST, port=PORT)








